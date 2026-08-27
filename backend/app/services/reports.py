"""Generacion de informes periodicos (Excel y PDF) por poligono.

Reune en una sola estructura (`ReportData`) todo lo que hace falta para
ambos formatos, para no duplicar las consultas: consumo y eficiencia por
nave en el periodo, alertas por severidad, incidencias por estado, y una
seleccion de las alertas mas relevantes. Los renderers (`render_excel`,
`render_pdf`) solo dan formato a esos datos.
"""

import io
from dataclasses import dataclass, field
from datetime import UTC, datetime, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.shapes import Drawing, String
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app import models
from app.services.efficiency import efficiency_scores, kwh_per_m2
from app.services.maintenance import maintenance_risk_score, risk_label
from app.services.timeseries import bucket_readings

PERIOD_LABELS = {"daily": "Diario", "weekly": "Semanal", "monthly": "Mensual"}
PERIOD_DELTAS = {
    "daily": timedelta(days=1),
    "weekly": timedelta(days=7),
    "monthly": timedelta(days=30),
}

# Estimaciones para dar contexto de negocio al consumo bruto. Son valores de
# referencia razonables para industria en España (precio medio de mercado
# libre y factor de emision de la red peninsular), no datos de facturacion
# real del cliente — el informe lo indica explicitamente.
ENERGY_PRICE_EUR_PER_KWH = 0.18
GRID_CO2_KG_PER_KWH = 0.21

BRAND_BLUE = "1D4ED8"
INK = "0F172A"
SLATE = "64748B"
CRITICAL = "DC2626"
WARNING = "D97706"
OK = "16A34A"


def _hex(value: str) -> colors.Color:
    """openpyxl usa hex sin '#'; reportlab lo exige. Los mismos literales
    (BRAND_BLUE, INK...) sirven para ambos formatos pasando por aqui."""
    return colors.HexColor(f"#{value}")


@dataclass
class BuildingReportRow:
    id: int
    code: str
    name: str
    building_type: str
    area_m2: float
    energy_kwh: float
    cost_eur: float
    efficiency_score: int
    alerts_critical: int
    alerts_warning: int
    maintenance_risk: int
    maintenance_label: str


@dataclass
class AlertRow:
    building_name: str
    severity: str
    message: str
    created_at: datetime
    status: str


@dataclass
class SeriesPoint:
    label: str
    value: float


@dataclass
class ReportData:
    polygon_name: str
    polygon_address: str | None
    period_key: str
    period_label: str
    period_start: datetime
    period_end: datetime
    generated_at: datetime
    total_energy_kwh: float
    energy_trend_pct: float | None
    avg_temperature: float | None
    critical_alerts: int
    warning_alerts: int
    resolved_alerts: int
    incidents_open: int
    incidents_in_progress: int
    incidents_resolved: int
    estimated_cost_eur: float
    estimated_co2_kg: float
    executive_summary: str
    buildings: list[BuildingReportRow] = field(default_factory=list)
    top_alerts: list[AlertRow] = field(default_factory=list)
    energy_series: list[SeriesPoint] = field(default_factory=list)


def _sensor_ids(db: Session, building_ids: list[int], sensor_type: models.SensorType) -> list[int]:
    if not building_ids:
        return []
    return db.execute(
        select(models.Sensor.id).where(
            models.Sensor.building_id.in_(building_ids),
            models.Sensor.sensor_type == sensor_type,
        )
    ).scalars().all()


def _energy_sum(db: Session, sensor_ids: list[int], since: datetime, until: datetime) -> float:
    if not sensor_ids:
        return 0.0
    return db.execute(
        select(func.coalesce(func.sum(models.SensorReading.value), 0.0)).where(
            models.SensorReading.sensor_id.in_(sensor_ids),
            models.SensorReading.timestamp >= since,
            models.SensorReading.timestamp < until,
        )
    ).scalar_one()


def build_report_data(db: Session, polygon: models.Polygon, period: str) -> ReportData:
    if period not in PERIOD_DELTAS:
        raise ValueError(f"Periodo desconocido: {period}")

    now = datetime.now(UTC)
    delta = PERIOD_DELTAS[period]
    since, until = now - delta, now
    previous_since = since - delta

    buildings = db.execute(
        select(models.Building).where(models.Building.polygon_id == polygon.id)
    ).scalars().all()
    building_ids = [b.id for b in buildings]

    energy_ids_by_building = {
        b.id: _sensor_ids(db, [b.id], models.SensorType.energy) for b in buildings
    }
    energy_by_building = {
        b.id: _energy_sum(db, energy_ids_by_building[b.id], since, until) for b in buildings
    }
    total_energy = sum(energy_by_building.values())

    all_energy_ids = _sensor_ids(db, building_ids, models.SensorType.energy)
    previous_total_energy = _energy_sum(db, all_energy_ids, previous_since, since)
    # Solo fiable si hay historico que cubra el periodo anterior entero: si
    # el dato mas antiguo es mas reciente que `previous_since`, esa ventana
    # esta incompleta y el porcentaje saldria disparado sin significar nada.
    earliest_reading = (
        db.execute(
            select(func.min(models.SensorReading.timestamp)).where(
                models.SensorReading.sensor_id.in_(all_energy_ids)
            )
        ).scalar_one()
        if all_energy_ids
        else None
    )
    # SQLite no guarda tz, así que vuelve como naive: se asume UTC, que es
    # lo unico que escribe la app (ver utcnow() en models.py).
    if earliest_reading is not None and earliest_reading.tzinfo is None:
        earliest_reading = earliest_reading.replace(tzinfo=UTC)
    has_full_previous_period = earliest_reading is not None and earliest_reading <= previous_since
    energy_trend_pct = (
        round((total_energy - previous_total_energy) / previous_total_energy * 100, 1)
        if has_full_previous_period and previous_total_energy > 0
        else None
    )

    temp_ids = _sensor_ids(db, building_ids, models.SensorType.temperature)
    avg_temperature = None
    if temp_ids:
        avg_temperature = db.execute(
            select(func.avg(models.SensorReading.value)).where(
                models.SensorReading.sensor_id.in_(temp_ids),
                models.SensorReading.timestamp >= since,
                models.SensorReading.timestamp < until,
            )
        ).scalar_one()
        avg_temperature = round(avg_temperature, 1) if avg_temperature is not None else None

    efficiency_values = {b.id: kwh_per_m2(energy_by_building[b.id], b.area_m2) for b in buildings}
    scores = efficiency_scores(efficiency_values)

    period_alerts = (
        db.execute(
            select(models.Alert).where(
                models.Alert.building_id.in_(building_ids),
                models.Alert.created_at >= since,
                models.Alert.created_at < until,
            )
        ).scalars().all()
        if building_ids
        else []
    )
    alerts_by_building: dict[int, list[models.Alert]] = {b.id: [] for b in buildings}
    for a in period_alerts:
        alerts_by_building.setdefault(a.building_id, []).append(a)

    critical_alerts = sum(1 for a in period_alerts if a.severity == models.AlertSeverity.critical)
    warning_alerts = sum(1 for a in period_alerts if a.severity == models.AlertSeverity.warning)
    resolved_alerts = sum(1 for a in period_alerts if a.status == models.AlertStatus.resolved)

    period_incidents = (
        db.execute(
            select(models.Incident).where(
                models.Incident.building_id.in_(building_ids),
                models.Incident.created_at >= since,
                models.Incident.created_at < until,
            )
        ).scalars().all()
        if building_ids
        else []
    )
    incidents_open = sum(1 for i in period_incidents if i.status == models.IncidentStatus.open)
    incidents_in_progress = sum(
        1 for i in period_incidents if i.status == models.IncidentStatus.in_progress
    )
    incidents_resolved = sum(
        1 for i in period_incidents if i.status == models.IncidentStatus.resolved
    )

    building_rows: list[BuildingReportRow] = []
    for b in buildings:
        b_alerts = alerts_by_building.get(b.id, [])
        b_critical = sum(1 for a in b_alerts if a.severity == models.AlertSeverity.critical)
        b_warning = sum(1 for a in b_alerts if a.severity == models.AlertSeverity.warning)
        risk = maintenance_risk_score(len(b_alerts), None, b.status.value)
        building_rows.append(
            BuildingReportRow(
                id=b.id,
                code=b.code,
                name=b.name,
                building_type=b.building_type,
                area_m2=b.area_m2,
                energy_kwh=round(energy_by_building[b.id], 2),
                cost_eur=round(energy_by_building[b.id] * ENERGY_PRICE_EUR_PER_KWH, 2),
                efficiency_score=scores[b.id],
                alerts_critical=b_critical,
                alerts_warning=b_warning,
                maintenance_risk=risk,
                maintenance_label=risk_label(risk),
            )
        )
    building_rows.sort(key=lambda r: r.energy_kwh, reverse=True)

    building_names = {b.id: b.name for b in buildings}
    top_alerts = [
        AlertRow(
            building_name=building_names.get(a.building_id, "?"),
            severity=a.severity.value,
            message=a.message,
            created_at=a.created_at,
            status=a.status.value,
        )
        for a in sorted(period_alerts, key=lambda a: a.created_at, reverse=True)[:20]
    ]

    # Serie para el grafico: por hora en el informe diario (24 puntos), por
    # dia en semanal/mensual (7 o ~30 puntos).
    granularity = "hour" if period == "daily" else "day"
    label_fmt = "%H:00" if period == "daily" else "%d/%m"
    energy_series: list[SeriesPoint] = []
    if all_energy_ids:
        rows = bucket_readings(
            db, all_energy_ids, since, granularity=granularity, agg="sum", until=until
        )
        for ts, value in rows:
            energy_series.append(SeriesPoint(label=ts.strftime(label_fmt), value=round(value, 2)))

    estimated_cost_eur = round(total_energy * ENERGY_PRICE_EUR_PER_KWH, 2)
    estimated_co2_kg = round(total_energy * GRID_CO2_KG_PER_KWH, 1)

    executive_summary = _build_executive_summary(
        total_energy_kwh=round(total_energy, 2),
        energy_trend_pct=energy_trend_pct,
        estimated_cost_eur=estimated_cost_eur,
        estimated_co2_kg=estimated_co2_kg,
        critical_alerts=critical_alerts,
        warning_alerts=warning_alerts,
        building_rows=building_rows,
        period_label=PERIOD_LABELS[period],
    )

    return ReportData(
        polygon_name=polygon.name,
        polygon_address=polygon.address,
        period_key=period,
        period_label=PERIOD_LABELS[period],
        period_start=since,
        period_end=until,
        generated_at=now,
        total_energy_kwh=round(total_energy, 2),
        energy_trend_pct=energy_trend_pct,
        avg_temperature=avg_temperature,
        critical_alerts=critical_alerts,
        warning_alerts=warning_alerts,
        resolved_alerts=resolved_alerts,
        incidents_open=incidents_open,
        incidents_in_progress=incidents_in_progress,
        incidents_resolved=incidents_resolved,
        estimated_cost_eur=estimated_cost_eur,
        estimated_co2_kg=estimated_co2_kg,
        executive_summary=executive_summary,
        buildings=building_rows,
        top_alerts=top_alerts,
        energy_series=energy_series,
    )


def _build_executive_summary(
    *,
    total_energy_kwh: float,
    energy_trend_pct: float | None,
    estimated_cost_eur: float,
    estimated_co2_kg: float,
    critical_alerts: int,
    warning_alerts: int,
    building_rows: list[BuildingReportRow],
    period_label: str,
) -> str:
    """Parrafo en lenguaje natural que sintetiza el periodo, al estilo del
    resumen ejecutivo de un informe corporativo: la cifra clave primero, el
    contexto (tendencia, coste, huella) despues, y una mencion nominal a la
    nave mejor y peor situada para que el lector sepa donde mirar."""
    if not building_rows:
        return (
            f"Este polígono todavía no tiene naves dadas de alta, así que no hay "
            f"datos de consumo que analizar en el periodo {period_label.lower()}."
        )

    trend_txt = ""
    if energy_trend_pct is not None:
        direction = "un aumento" if energy_trend_pct > 0 else "una reducción"
        trend_txt = (
            f", lo que supone {direction} del {abs(energy_trend_pct):.1f}% "
            f"respecto al periodo anterior"
        )

    best = max(building_rows, key=lambda b: b.efficiency_score)
    highest_risk = max(building_rows, key=lambda b: b.maintenance_risk)

    if not critical_alerts and not warning_alerts:
        alerts_txt = "sin alertas activas destacables"
    else:
        alerts_txt = (
            f"{critical_alerts} alerta{'s' if critical_alerts != 1 else ''} crítica"
            f"{'s' if critical_alerts != 1 else ''} y {warning_alerts} de aviso"
        )

    risk_txt = ""
    if highest_risk.maintenance_risk >= 35:
        risk_txt = (
            f" La nave con mayor riesgo de mantenimiento es {highest_risk.code} "
            f"({highest_risk.maintenance_label.lower()}, {highest_risk.maintenance_risk}/100)."
        )

    energy_fmt = f"{total_energy_kwh:,.0f}".replace(",", ".")
    cost_fmt = f"{estimated_cost_eur:,.0f}".replace(",", ".")
    co2_fmt = f"{estimated_co2_kg:,.0f}".replace(",", ".")

    return (
        f"El polígono consumió {energy_fmt} kWh en el periodo {period_label.lower()}"
        f"{trend_txt}, con un coste estimado de {cost_fmt} € y una huella de "
        f"{co2_fmt} kg de CO2 equivalente. Se registraron {alerts_txt}. "
        f"La nave más eficiente fue {best.code} — {best.name.split(' - ')[-1]} "
        f"(puntuación {best.efficiency_score}/100).{risk_txt}"
    )


def _fmt_dt(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y %H:%M")


def _fmt_date(dt: datetime) -> str:
    return dt.strftime("%d/%m/%Y")


# ---------------------------------------------------------------- Excel ----

def render_excel(report: ReportData) -> bytes:
    wb = Workbook()

    header_fill = PatternFill("solid", fgColor=BRAND_BLUE)
    header_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=16, color=INK)
    subtitle_font = Font(size=10, color=SLATE)
    label_font = Font(bold=True, size=10, color=SLATE)
    value_font = Font(bold=True, size=14, color=INK)
    thin = Side(style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header_row(ws, row: int, ncols: int):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def autosize(ws, widths: list[int]):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # --- Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Informe {report.period_label.lower()} — {report.polygon_name}"
    ws["A1"].font = title_font
    ws.merge_cells("A2:D2")
    ws["A2"] = (
        f"{_fmt_date(report.period_start)} — {_fmt_date(report.period_end)}"
        f"  ·  generado el {_fmt_dt(report.generated_at)}"
    )
    ws["A2"].font = subtitle_font
    if report.polygon_address:
        ws.merge_cells("A3:D3")
        ws["A3"] = report.polygon_address
        ws["A3"].font = subtitle_font

    kpis = [
        ("Consumo total", f"{report.total_energy_kwh:,.1f} kWh".replace(",", ".")),
        (
            "Tendencia vs. periodo anterior",
            f"{report.energy_trend_pct:+.1f}%" if report.energy_trend_pct is not None else "—",
        ),
        (
            "Temperatura media",
            f"{report.avg_temperature:.1f} °C" if report.avg_temperature is not None else "—",
        ),
        ("Naves", str(len(report.buildings))),
        ("Coste estimado", f"{report.estimated_cost_eur:,.0f} €".replace(",", ".")),
        ("Huella de CO₂ estimada", f"{report.estimated_co2_kg:,.0f} kg".replace(",", ".")),
        ("Alertas críticas", str(report.critical_alerts)),
        ("Alertas de aviso", str(report.warning_alerts)),
        ("Incidencias abiertas", str(report.incidents_open)),
        ("Incidencias resueltas", str(report.incidents_resolved)),
    ]
    row = 5
    for i, (label, value) in enumerate(kpis):
        r = row + (i // 2)
        c = 1 + (i % 2) * 2
        ws.cell(row=r, column=c, value=label).font = label_font
        ws.cell(row=r, column=c + 1, value=value).font = value_font
    autosize(ws, [24, 16, 24, 16])

    summary_row = row + (len(kpis) // 2) + 2
    ws.cell(row=summary_row, column=1, value="Resumen ejecutivo").font = Font(
        bold=True, size=11, color=INK
    )
    ws.merge_cells(start_row=summary_row + 1, start_column=1, end_row=summary_row + 4, end_column=4)
    summary_cell = ws.cell(row=summary_row + 1, column=1, value=report.executive_summary)
    summary_cell.font = Font(size=10, color=INK)
    summary_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Naves ---
    ws2 = wb.create_sheet("Naves")
    headers = [
        "Código",
        "Nave",
        "Tipo",
        "Superficie (m²)",
        "Consumo (kWh)",
        "Coste estimado (€)",
        "Eficiencia (0-100)",
        "Alertas críticas",
        "Alertas aviso",
        "Riesgo mantenimiento",
    ]
    for c, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=c, value=h)
    style_header_row(ws2, 1, len(headers))
    for r, b in enumerate(report.buildings, start=2):
        values = [
            b.code,
            b.name,
            b.building_type,
            b.area_m2,
            b.energy_kwh,
            b.cost_eur,
            b.efficiency_score,
            b.alerts_critical,
            b.alerts_warning,
            f"{b.maintenance_label} ({b.maintenance_risk})",
        ]
        for c, v in enumerate(values, start=1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = border
            if b.alerts_critical > 0 and c == 8:
                cell.font = Font(color=CRITICAL, bold=True)
    autosize(ws2, [8, 28, 14, 14, 14, 16, 16, 14, 12, 20])

    if report.buildings:
        n = len(report.buildings)
        chart = BarChart()
        chart.type = "col"
        chart.title = "Consumo por nave (kWh)"
        chart.y_axis.title = "kWh"
        chart.style = 10
        chart.height = 8
        chart.width = 18
        data = Reference(ws2, min_col=5, min_row=1, max_row=n + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=n + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = False
        series = chart.series[0]
        series.graphicalProperties.solidFill = BRAND_BLUE
        ws2.add_chart(chart, f"A{n + 4}")

    # --- Alertas ---
    ws3 = wb.create_sheet("Alertas")
    headers3 = ["Fecha", "Nave", "Severidad", "Mensaje", "Estado"]
    for c, h in enumerate(headers3, start=1):
        ws3.cell(row=1, column=c, value=h)
    style_header_row(ws3, 1, len(headers3))
    for r, a in enumerate(report.top_alerts, start=2):
        values = [
            _fmt_dt(a.created_at),
            a.building_name,
            "Crítica" if a.severity == "critical" else "Aviso",
            a.message,
            "Resuelta" if a.status == "resolved" else "Activa",
        ]
        for c, v in enumerate(values, start=1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = border
            if c == 3:
                cell.font = Font(color=CRITICAL if a.severity == "critical" else WARNING, bold=True)
    autosize(ws3, [18, 28, 12, 60, 12])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------ PDF ----

def render_pdf(report: ReportData) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        topMargin=18 * mm,
        bottomMargin=16 * mm,
        leftMargin=16 * mm,
        rightMargin=16 * mm,
        title=f"Informe {report.period_label} — {report.polygon_name}",
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleINDU", parent=styles["Title"], fontSize=20, textColor=_hex(INK),
        spaceAfter=2,
    )
    subtitle_style = ParagraphStyle(
        "SubtitleINDU", parent=styles["Normal"], fontSize=10, textColor=_hex(SLATE),
    )
    section_style = ParagraphStyle(
        "SectionINDU", parent=styles["Heading2"], fontSize=13, textColor=_hex(INK),
        spaceBefore=14, spaceAfter=6,
    )
    body_style = ParagraphStyle("BodyINDU", parent=styles["Normal"], fontSize=9, leading=13)

    elements = []

    elements.append(Paragraph("INDU-TWIN", ParagraphStyle(
        "Brand", parent=styles["Normal"], fontSize=9, textColor=_hex(BRAND_BLUE),
        fontName="Helvetica-Bold",
    )))
    elements.append(Paragraph(f"Informe {report.period_label.lower()}", title_style))
    elements.append(Paragraph(
        f"{report.polygon_name}"
        + (f" · {report.polygon_address}" if report.polygon_address else ""),
        subtitle_style,
    ))
    elements.append(Paragraph(
        f"Periodo: {_fmt_date(report.period_start)} – {_fmt_date(report.period_end)}"
        f"  ·  Generado el {_fmt_dt(report.generated_at)}",
        subtitle_style,
    ))
    elements.append(Spacer(1, 6))
    elements.append(HRFlowable(width="100%", thickness=1.2, color=_hex(BRAND_BLUE)))
    elements.append(Spacer(1, 12))

    # --- KPI cards (as a table so they lay out in a tidy grid) ---
    def kpi_cell(label: str, value: str, accent: str = INK) -> Table:
        t = Table(
            [[Paragraph(label, ParagraphStyle(
                "kpiLabel", fontSize=8, textColor=_hex(SLATE), fontName="Helvetica",
            ))], [Paragraph(value, ParagraphStyle(
                "kpiValue", fontSize=15, textColor=_hex(accent),
                fontName="Helvetica-Bold", spaceBefore=2,
            ))]],
            colWidths=[42 * mm],
        )
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), _hex("F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.6, _hex("E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ]))
        return t

    trend_txt = f"{report.energy_trend_pct:+.1f}%" if report.energy_trend_pct is not None else "—"
    trend_color = CRITICAL if (report.energy_trend_pct or 0) > 15 else INK
    temp_txt = f"{report.avg_temperature:.1f} °C" if report.avg_temperature is not None else "—"
    cost_txt = f"{report.estimated_cost_eur:,.0f} €".replace(",", ".")
    co2_txt = f"{report.estimated_co2_kg:,.0f} kg".replace(",", ".")

    critical_color = CRITICAL if report.critical_alerts else INK
    warning_color = WARNING if report.warning_alerts else INK
    open_incidents = report.incidents_open + report.incidents_in_progress

    blank_cell = Table([[""]], colWidths=[44 * mm])
    blank_cell.setStyle(TableStyle([("BOX", (0, 0), (-1, -1), 0, colors.white)]))

    kpi_row1 = [
        kpi_cell("CONSUMO TOTAL", f"{report.total_energy_kwh:,.0f} kWh".replace(",", ".")),
        kpi_cell("VS. PERIODO ANTERIOR", trend_txt, trend_color),
        kpi_cell("COSTE ESTIMADO", cost_txt),
        kpi_cell("HUELLA DE CO2 ESTIMADA", co2_txt),
    ]
    kpi_row2 = [
        kpi_cell("TEMPERATURA MEDIA", temp_txt),
        kpi_cell("NAVES", str(len(report.buildings))),
        kpi_cell("ALERTAS CRÍTICAS", str(report.critical_alerts), critical_color),
        kpi_cell("ALERTAS DE AVISO", str(report.warning_alerts), warning_color),
    ]
    kpi_row3 = [
        kpi_cell("INCIDENCIAS ABIERTAS", str(open_incidents)),
        kpi_cell("INCIDENCIAS RESUELTAS", str(report.incidents_resolved), OK),
        blank_cell,
        blank_cell,
    ]
    kpi_table = Table([kpi_row1, kpi_row2, kpi_row3], colWidths=[44 * mm] * 4, hAlign="LEFT")
    kpi_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 3))
    elements.append(Paragraph(
        f"Coste y CO2 estimados a {ENERGY_PRICE_EUR_PER_KWH:.2f} €/kWh y "
        f"{GRID_CO2_KG_PER_KWH:.2f} kg CO2/kWh (mix medio de red peninsular) — no proceden de "
        f"una factura real.",
        ParagraphStyle("Disclaimer", parent=body_style, fontSize=7, textColor=_hex(SLATE)),
    ))

    # --- Resumen ejecutivo ---
    elements.append(Paragraph("Resumen ejecutivo", section_style))
    summary_box = Table(
        [[Paragraph(report.executive_summary, ParagraphStyle(
            "ExecSummary", parent=body_style, fontSize=9.5, leading=14, textColor=_hex(INK),
        ))]],
        colWidths=[178 * mm],
    )
    summary_box.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), _hex("F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.6, _hex(BRAND_BLUE)),
        ("LINEBEFORE", (0, 0), (0, 0), 3, _hex(BRAND_BLUE)),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 12),
        ("RIGHTPADDING", (0, 0), (-1, -1), 12),
    ]))
    elements.append(summary_box)

    # --- Gráfico de consumo por nave ---
    if report.buildings:
        elements.append(Paragraph("Consumo por nave", section_style))
        chart_values = [b.energy_kwh for b in report.buildings]
        chart_labels = [b.code for b in report.buildings]
        max_val = max(chart_values) if chart_values else 1

        drawing = Drawing(178 * mm, 55 * mm)
        bar = VerticalBarChart()
        bar.x = 30
        bar.y = 15
        bar.width = 178 * mm - 50
        bar.height = 55 * mm - 30
        bar.data = [chart_values]
        bar.categoryAxis.categoryNames = chart_labels
        bar.categoryAxis.labels.fontSize = 8
        bar.categoryAxis.labels.fillColor = _hex(SLATE)
        bar.valueAxis.valueMin = 0
        bar.valueAxis.valueMax = max_val * 1.15 if max_val else 1
        bar.valueAxis.labels.fontSize = 7
        bar.valueAxis.labels.fillColor = _hex(SLATE)
        bar.bars[0].fillColor = _hex(BRAND_BLUE)
        bar.barWidth = 12
        bar.groupSpacing = 6
        bar.strokeColor = None
        drawing.add(bar)
        drawing.add(String(0, 55 * mm - 8, "kWh", fontSize=7, fillColor=_hex(SLATE)))
        elements.append(drawing)

    # --- Ranking de naves ---
    elements.append(Paragraph("Consumo, coste y eficiencia por nave", section_style))
    if report.buildings:
        rows = [["Nave", "Tipo", "Consumo", "Coste", "Eficiencia", "Alertas", "Riesgo mant."]]
        for b in report.buildings:
            alerts_txt = f"{b.alerts_critical} crít. / {b.alerts_warning} aviso"
            rows.append([
                Paragraph(f"{b.code} · {b.name}", body_style),
                b.building_type,
                f"{b.energy_kwh:,.1f} kWh".replace(",", "."),
                f"{b.cost_eur:,.0f} €".replace(",", "."),
                str(b.efficiency_score),
                alerts_txt,
                f"{b.maintenance_label} ({b.maintenance_risk})",
            ])
        col_widths = [44 * mm, 18 * mm, 22 * mm, 18 * mm, 18 * mm, 26 * mm, 30 * mm]
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), _hex(BRAND_BLUE)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _hex("F8FAFC")]),
            ("GRID", (0, 0), (-1, -1), 0.4, _hex("E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]
        for i, b in enumerate(report.buildings, start=1):
            if b.alerts_critical > 0:
                style.append(("TEXTCOLOR", (5, i), (5, i), _hex(CRITICAL)))
        t.setStyle(TableStyle(style))
        elements.append(t)
    else:
        elements.append(Paragraph("Este polígono todavía no tiene naves.", body_style))

    # --- Alertas destacadas ---
    elements.append(Paragraph("Alertas del periodo", section_style))
    if report.top_alerts:
        severity_style = {
            "critical": ParagraphStyle(
                "sevCrit", parent=body_style, textColor=_hex(CRITICAL), fontName="Helvetica-Bold",
            ),
            "warning": ParagraphStyle(
                "sevWarn", parent=body_style, textColor=_hex(WARNING), fontName="Helvetica-Bold",
            ),
        }
        rows = [["Fecha", "Nave", "Severidad", "Mensaje"]]
        for a in report.top_alerts:
            rows.append([
                a.created_at.strftime("%d/%m %H:%M"),
                Paragraph(a.building_name, body_style),
                Paragraph(
                    "Crítica" if a.severity == "critical" else "Aviso",
                    severity_style[a.severity],
                ),
                Paragraph(a.message, body_style),
            ])
        t = Table(rows, colWidths=[20 * mm, 38 * mm, 18 * mm, 98 * mm], repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), _hex(INK)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("GRID", (0, 0), (-1, -1), 0.4, _hex("E2E8F0")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        t.setStyle(TableStyle(style))
        elements.append(t)
    else:
        elements.append(Paragraph("Sin alertas en este periodo.", body_style))

    def _footer(canvas, _doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(_hex(SLATE))
        canvas.drawString(16 * mm, 10 * mm, "INDU-TWIN · Digital Twin SaaS")
        canvas.drawRightString(
            A4[0] - 16 * mm, 10 * mm, f"Página {canvas.getPageNumber()}"
        )
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()
